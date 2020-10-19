# -*- coding=utf-8 -*-
import requests
import xlsxwriter
import datetime
import xlrd
import openpyxl
from openpyxl.styles import Font, Border, Side, PatternFill, colors, Alignment

from bs4 import BeautifulSoup
url='http://top.baidu.com/buzz?b=1&fr=20811'
headers= {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/55.0.2883.87 Safari/537.36'}
file='D:\\项目\\PythonApplication1\\百度热搜.xlsx'
workbook=openpyxl.load_workbook(file)
workbook.guess_type=True
worksheet=workbook.active



worksheet['A2']="序号"
worksheet['B2']="内容"
worksheet['C2']="热度"
m=1


html=requests.get(url,headers=headers)
page=html.text.encode('iso-8859-1').decode('gbk')
#print(page)
soup=BeautifulSoup(page,'html.parser')
tops=soup.find_all('a','list-title')
hots=soup.find_all('td','last')

i=1

while worksheet['A{}'.format(m)].value!=None:
	m=m+1

worksheet.merge_cells('A{}:C{}'.format(str(m),str(m)))
now=datetime.datetime.now().strftime('%Y-%m-%d  %H')
worksheet['A{}'.format(str(m))]=now+":00 的百度热搜"
worksheet.column_dimensions['B'].width=40
worksheet.row_dimensions[m].height=30
worksheet['A{}'.format(str(m))].font = Font(name='Times New Roman', size=16, bold=True, italic=True, color=colors.RED)
worksheet['A{}'.format(str(m))].alignment = Alignment(horizontal='center',vertical='center')

for top,hot in zip(tops,hots):
	m=m+1
	worksheet['A{}'.format(m)]=i
	worksheet['B{}'.format(m)]=top.text
	worksheet['C{}'.format(m)]=hot.text
	print(str(i)+"  "+top.text+"  "+hot.text)
	i=i+1
	
workbook.save(file)

